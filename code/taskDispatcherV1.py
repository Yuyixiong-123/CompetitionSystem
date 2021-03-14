# -*- coding: utf-8 -*-
"""
Created on Mon Feb  8 17:03:50 2021

在该界面就相当于是一个class 指挥者，它可以调用全部的city和resource对象，来达到没有飞机空闲的目的
"""
# =============================================================================
# =============================================================================
# # V1.1版本更新说明：随机选择城市，而非最近的城市，扩展了可能性，能够在更大样本中寻找最优值。
# =============================================================================
# =============================================================================
# V2版本更新，完善了信息输出机制，能够将 任一方案 的仿真结果以图表和log txt的形式储存起来
# =============================================================================
# =============================================================================
# V3版本更新，调整为从时间权重的角度计算方案得分，而不再是“仅考虑任务完成总时长”
# =============================================================================
# buglog
# 飞机执行功能时没有考虑resource剩余的量，默认是充足的。
# 实际上机场不能同时保障10来架AC311，也不可能让他们同时起飞，模型中却是如此操作的（调度中手速也跟不上）
# 用飞机作业的总时间来计算油量，而不是不计算作业时间，只计算飞行时间;同理在检查油量是否足够的时候也修改补充
# 检查运算的单位，小时和秒的统一，时间是按照秒s来算的，而油耗，距离都是是按照小时算的
# 修改变量名称和方法名称
# 整理侦察时的油耗模型
# 模型的随机性在于安排飞机去随机一个城市的随机一种任务
# 存在最后一次运走的飞机结束时间，并非该任务最终结束的时间
# =============================================================================
import openpyxl
import math
from math import  cos,sin
import random
import matplotlib.pyplot as plt


plan='.\\allocationPlan\\plan2'
realTime=0
realTimeList=[]
loglist=[]
taskDoneTime=[0,0,0,0,0,0,0,0,0]
taskWeightList=[0.15, 0.05, 0.05, 0.16, 0.06, 0.05, 0.16, 0.17, 0.15]
taskDoneHeli=['','','','','','','','','']
# cityLock=[True,True,True,True,True,True,True,True,True,]
d0={'name':[],'t':[]}
d1={'name':[],'t':[]}
d2={'name':[],'t':[]}
d3={'name':[],'t':[]}
d4={'name':[],'t':[]}
d5={'name':[],'t':[]}
d6={'name':[],'t':[]}
d7={'name':[],'t':[]}
d8={'name':[],'t':[]}
register={0:d0,1:d1,2:d2,3:d3,4:d4,5:d5,6:d6,7:d7,8:d8}


def checkTaskHaveDone(t,name,cityloc,cityneed):#输入当前直升机完成该任务的时间，并在该任务完成了之后，将此时间作为任务被完成的时间，记录下来
     taskLoc=['建德','建德','丽水','丽水','新昌','新昌','缙云','缙云','乐清']     
     taskneed=["equipIn", "peopleIn", "loadIn", "peopleOut", "loadIn", "woundOut", "peopleOut", "scout", "peopleOut"]
     #针对每一个任务，都定义一个二维（2行 n列）字典，第一行表示登记的飞机名称，第二行表示登记的飞机作业时间
     
     for i in range(len(taskLoc)):
          if cityloc==taskLoc[i] and cityneed==taskneed[i]:#此时可以判定，飞机是在执行第i+1个任务，需要记录该飞机名称及其时间
               register[i]['name'].append(name)
               register[i]['t'].append(t)

def taskSettle():
     for i in range(len(taskDoneTime)):
          taskDoneTime[i]=max(register[i]['t'])
          taskDoneHeli[i]=register[i]['name'][register[i]['t'].index(taskDoneTime[i])]

def rad(d):
    return d * math.pi / 180.0

def getDistance(lat1, lng1, lat2, lng2):
    EARTH_REDIUS = 6378.137
    radLat1 = rad(lat1)
    radLat2 = rad(lat2)
    a = radLat1 - radLat2
    b = rad(lng1) - rad(lng2)
    s = 2 * math.asin(math.sqrt(math.pow(sin(a/2), 2) + cos(radLat1) * cos(radLat2) * math.pow(sin(b/2), 2)))
    s = s * EARTH_REDIUS
    return s#计算结果单位为千米(km)

def getMaxOilCapa(name):
     heliNameList=["Mi-26", "AC313基本型", "S-76", "H155", "AC312", "AC311", "Bell429", "H135医疗型", "Mi-171", "H225", "长鹰5E", "AC313医疗型"]
     oilMaxList=[12000, 3500, 818, 993, 580, 423, 600, 560, 1700, 2044, 40, 3500]
     for i in range(len(heliNameList)):
          if heliNameList[i] in name:
               return oilMaxList[i]

class Heli:
     t=0
    
     def __init__(self, name, lon, lat, price, speed, oilMass, oilConsumeSpeed, carryPeople, peopleUpDownTime, load, loadUpDownTime, scoutSpeed, carryWounded, woundedUpDownTime, equipUpDown, equipUpDownTime):
          self.name=name
          self.lon=lon
          self.lat=lat
          self.price=price
          self.speed=speed
          self.oilMass=oilMass
          self.oilConsumeSpeed=oilConsumeSpeed
          self.carryPeople=carryPeople
          self.peopleUpDownTime=peopleUpDownTime
          self.load=load
          self.loadUpDownTime=loadUpDownTime
          self.scoutSpeed=scoutSpeed
          self.carryWounded=carryWounded
          self.woundedUpDownTime=woundedUpDownTime
          self.equipUpDown=equipUpDown
          self.equipUpDownTime=equipUpDownTime
          
     # 接下来依次定义了飞机做某一项任务的时候对全局信息的变更,但是仅对飞机和需求城市进行了变更，因为程序是基于需求的，而资源总是充足的。
     # 注意其中的油量消耗，是按照h为单位乘以速率的，所以需要转化单位//其中航速，距离正好是小时单位，所以不考虑，主要是侦察功能，需要统一
     def exepeopleIn(self,city,resource):#运入人员

          #when the heli is sure to do the mission ,use this function to change the fields
          taskLoad=min(city.peopleIn , self.carryPeople)
          d=getDistance(self.lat, self.lon,resource.lat,resource.lon)+getDistance(resource.lat,resource.lon,  city.lat, city.lon)
         
          #calculate the fields of city and helicopter ,change them to sync 
          d_oilCheck=d+getDistance(city.lat, city.lon, 29.34, 120.03)
          if (self.oilMass-self.oilConsumeSpeed*(d_oilCheck/self.speed+2*taskLoad*self.peopleUpDownTime/3600)>0):
               city.peopleIn-=taskLoad
             
               self.oilMass-=self.oilConsumeSpeed*(d/self.speed+2*taskLoad*self.peopleUpDownTime/3600)
               self.lat=city.lat
               self.lon=city.lon
               self.t+=d/self.speed*3600
               self.t+=2*taskLoad*self.peopleUpDownTime  #up and down so double that
               # print(self.name,'  exepeopleIn  ',taskLoad, city.name, resource.name,'剩余需求',city.peopleIn)
               loglist.append([self.name,'  exepeopleIn  ',taskLoad, city.name, resource.name,'剩余需求',city.peopleIn,'飞机油量：',self.oilMass,'飞机作业时间：',self.t])
               checkTaskHaveDone(self.t,self.name,city.name,'peopleIn')
          else:
               self.backToBase()#如果油量判断是不够的话，就飞回基地去了，因为都是就近原则，所以大概率上符合实际，不用重新规划任务
          
     def exepeopleOut(self,city,resource):#转移人员
          taskLoad=min(city.peopleOut,self.carryPeople)
          d=getDistance(self.lat, self.lon,city.lat, city.lon)+getDistance(resource.lat,resource.lon,  city.lat, city.lon)

          d_oilCheck=d+getDistance(resource.lat,resource.lon,29.34, 120.03)
          if (self.oilMass-self.oilConsumeSpeed*(d_oilCheck/self.speed+2*taskLoad*self.peopleUpDownTime/3600)>0):
               city.peopleOut-=taskLoad
               self.oilMass-=self.oilConsumeSpeed*(d/self.speed+2*taskLoad*self.peopleUpDownTime/3600)
               self.lat=resource.lat
               self.lon=resource.lon
               self.t+=d/self.speed*3600
               self.t+=2*taskLoad*self.peopleUpDownTime  #up and down so double that
               # print(self.name,'  exepeopleOut  ',taskLoad, city.name, resource.name,'剩余需求',city.peopleOut)
               loglist.append([self.name,'  exepeopleOut  ',taskLoad, city.name, resource.name,'剩余需求',city.peopleOut,'飞机油量：',self.oilMass,'飞机作业时间：',self.t])
               checkTaskHaveDone(self.t,self.name,city.name,'peopleOut')
          else:
               self.backToBase()

     def exeloadIn(self,city,resource):#去r得到物资，然后运到city
          taskLoad=min(city.loadIn,self.load)
          d=getDistance(self.lat, self.lon,resource.lat,resource.lon)+getDistance(resource.lat,resource.lon,  city.lat, city.lon)

          d_oilCheck=d+getDistance(city.lat, city.lon, 29.34, 120.03)
          if (self.oilMass-self.oilConsumeSpeed*(d_oilCheck/self.speed+2*taskLoad*self.loadUpDownTime/3600)>0):
               city.loadIn-=taskLoad
               self.oilMass-=self.oilConsumeSpeed*(d/self.speed+2*taskLoad*self.loadUpDownTime/3600)
               self.lat=city.lat
               self.lon=city.lon
               self.t+=d/self.speed*3600
               self.t+=2*taskLoad*self.loadUpDownTime  #up and down so double that
               # print(self.name,'  exeloadIn  ',taskLoad, city.name, resource.name,'剩余需求',city.loadIn)
               loglist.append([self.name,'  exeloadIn  ',taskLoad, city.name, resource.name,'剩余需求',city.loadIn,'飞机油量：',self.oilMass,'飞机作业时间：',self.t])
               checkTaskHaveDone(self.t,self.name,city.name,'loadIn')
          else:
               self.backToBase()#如果油量判断是不够的话，就飞回基地去了，因为都是就近原则，所以大概率上符合实际，不用重新规划任务

     def exeequipIn(self,city,resource):#去r得到设备，然后运到city
          taskLoad=1
          d=getDistance(self.lat, self.lon,resource.lat,resource.lon)+getDistance(resource.lat,resource.lon,  city.lat, city.lon)
          
          d_oilCheck=d+getDistance(city.lat, city.lon, 29.34, 120.03)
          if (self.oilMass-self.oilConsumeSpeed*(d_oilCheck/self.speed+2*taskLoad*self.equipUpDownTime/3600)>0):
               city.equipIn-=taskLoad
               self.oilMass-=self.oilConsumeSpeed*(d/self.speed+2*taskLoad*self.equipUpDownTime/3600)
               self.lat=city.lat
               self.lon=city.lon
               self.t+=d/self.speed*3600
               self.t+=2*taskLoad*self.equipUpDownTime #up and down so double that
               # print(self.name,'  exeequipIn  ',taskLoad, city.name, resource.name,'剩余需求',city.equipIn)
               loglist.append([self.name,'  exeequipIn  ',taskLoad, city.name, resource.name,'剩余需求',0,'飞机油量：',self.oilMass,'飞机作业时间：',self.t])
               checkTaskHaveDone(self.t,self.name,city.name,'equipIn')
          else:
               self.backToBase()#如果油量判断是不够的话，就飞回基地去了，因为都是就近原则，所以大概率上符合实际，不用重新规划任务
     
     def exewoundOut(self,city,resource):#去city得到重伤员，然后运到医院resource
          taskLoad=min(city.woundOut,self.carryWounded)
          d=getDistance(self.lat, self.lon,city.lat, city.lon)+getDistance(resource.lat,resource.lon,  city.lat, city.lon)
         
          d_oilCheck=d+getDistance(resource.lat,resource.lon,29.34, 120.03)
          if (self.oilMass-self.oilConsumeSpeed*(d_oilCheck/self.speed+2*taskLoad*self.woundedUpDownTime/3600)>0):
               city.woundOut-=taskLoad
               self.oilMass-=self.oilConsumeSpeed*(d/self.speed+2*taskLoad*self.woundedUpDownTime/3600)
               self.lat=city.lat
               self.lon=city.lon
               self.t+=d/self.speed*3600
               self.t+=2*taskLoad*self.woundedUpDownTime  #up and down so double that
               # print(self.name,'  exewoundOut  ',taskLoad, city.name, resource.name,'剩余需求',city.woundOut)
               loglist.append([self.name,'  exewoundOut  ',taskLoad, city.name, resource.name,'剩余需求',city.woundOut,'飞机油量：',self.oilMass,'飞机作业时间：',self.t])
               checkTaskHaveDone(self.t,self.name,city.name,'woundOut')
          else:
               self.backToBase()
          
     def exescout(self,city,resource):#直接去城市city侦察然后呆在那就好
          taskLoad=city.scout#这里是有bug的，没有考虑飞机的油量,就直接认为可以一次侦察完了，主要是考虑到直升机过去侦察，一般也就30分钟
          d=getDistance(self.lat, self.lon,city.lat, city.lon)
          
          d_oilCheck=d+getDistance(city.lat, city.lon, 29.34, 120.03)
          if (self.oilMass-self.oilConsumeSpeed*(d_oilCheck/self.speed+taskLoad*self.scoutSpeed/3600)>0):
               city.scout=0
               self.oilMass-=self.oilConsumeSpeed*(d/self.speed+taskLoad*self.scoutSpeed/3600)
               self.lat=city.lat
               self.lon=city.lon
               self.t+=d/self.speed*3600
               self.t+=taskLoad*self.scoutSpeed
               # print(self.name,'  exescout  ',taskLoad, city.name, resource.name,'剩余需求',city.scout)
               loglist.append([self.name,'  exescout  ',taskLoad, city.name, resource.name,'剩余需求',city.scout,'飞机油量：',self.oilMass,'飞机作业时间：',self.t])
               checkTaskHaveDone(self.t,self.name,city.name,'scout')
          
     #在做各任务的过程中，首先判断一下油量是不是足够，（因为每种任务的distance前后顺序不一样，所以放入每一种任务里面
     #如果油量不够的话，就调用“返回基地”子函数
     
     def exeFireExtinct(self,city,resource):
         pass 
         # confirm the scout task has been done，if done then do, if not then pass, make sure the execution log be deleted
          # 
          
     
     def backToBase(self):
          # 注意这里在简单版本下可以直接用单个基地，如果有多个基地就要选择。则传入参数resourceList,查看那个基地可以保障，然后选择最近的
          d=getDistance(self.lat, self.lon, 29.34, 120.03)
          # self.oilMass-=self.oilConsumeSpeed*d/self.speed  注意程序中缺少一个“油量是否够回机场的判断”
          # 而且，没有计算“所需加油量”，直接用20分钟保障时间了，本身应该根据加油速度有所不同
          # print(self.name,'  backToBase  剩余油量：',self.oilMass,'只能飞行：',self.oilMass/self.oilConsumeSpeed,'h')
          self.t+=d/self.speed
          self.oilMass=getMaxOilCapa(self.name)
          self.t+=1200
          self.lat=29.34
          self.lon=120.03
          loglist.append([self.name,'  backToBase  ', '飞机作业时间：',self.t])
     
     
     
class City:
     def __init__(self,name, lon, lat, peopleIn, peopleOut, loadIn, equipIn, scout, woundOut):
          self.name=name
          self.lon=lon
          self.lat=lat
          self.peopleIn=peopleIn
          self.peopleOut=peopleOut
          self.loadIn=loadIn
          self.equipIn=equipIn
          self.scout=scout
          self.woundOut=woundOut
     
     def checkNeeded(self):
          #调用此方法，能够返回该城市需求所对应的“所需要能力”的名称，用于给飞机找任务时，观察“能不能为这个城市做点什么”
          # 返回的列表中  如果有飞机也有的能力，就可以将该城市，以及需求类型 作为输入变量
          needList=[]
          
          if self.equipIn==0:
               pass
          else:
               needList.append('equipIn')
               
          if self.peopleIn==0:
               pass
          else:
               needList.append('peopleIn')
               
          if self.peopleOut==0:
               pass
          else:
               needList.append('peopleOut')
               
          if self.loadIn==0:
               pass
          else:
               needList.append('loadIn')
          
          if self.scout==0:
               pass
          else:
               needList.append('scout')
          
          if self.woundOut==0:
               pass
          else:
               needList.append('woundOut') 
          
          return needList#返回了当前城市还存在的需求

          

class Resource:
     def __init__(self,name, lon, lat, refuelTime, peopleCapacity , reliefWorker, load, equip , woundedCapacity):
          self.name=name
          self.lon=lon
          self.lat=lat
          self.refuelTime=refuelTime
          self.peopleCapacity =peopleCapacity 
          self.reliefWorker=reliefWorker
          self.load=load
          self.equip =equip 
          self.woundedCapacity=woundedCapacity



def findTheNearestCity(lat,lon,locationList):
     #根据传入的城市列表和坐标点，输出离坐标点最近的城市
     dmin=100000000000
     loc=None
     for location in locationList:
          d=getDistance(lat, lon, location.lat, location.lon)
          if(d<dmin):
               dmin=d
               loc=location
     return loc

def findRandomCity(lat,lon,locationList):
     return random.choice(locationList)

def findMatchResourceList(need,resourceList):
     # 根据城市所需要的需求，找到能够匹配相关需求的资源城市
     rList=[]
     if(need=='peopleIn'):
          for r in resourceList:
               if(r.reliefWorker>0):
                    rList.append(r)
     if(need=='peopleOut'):
          for r in resourceList:
               if(r.peopleCapacity >0):
                    rList.append(r)               
     if(need=='loadIn'):
          for r in resourceList:
               if(r.load>0):
                    rList.append(r)
     if(need=='woundOut'):
          for r in resourceList:
               if(r.woundedCapacity>0):
                    rList.append(r)
     if(need=='equipIn'):
          for r in resourceList:
               if(r.equip>0):
                    rList.append(r)
     if(need=='scout'):
          rList.append(resourceList[0])
     return rList
     
def doOneFunctionByNeeded(heli,need,city,resource):
     # 根据传入的直升机，任务类型，以及提前指定好的城市/资源，指派直升机执行相应的功能
     if(need=='peopleIn'):
          heli.exepeopleIn(city,resource)
     if(need=='peopleOut'):
          heli.exepeopleOut(city,resource)
     if(need=='scout'):
          heli.exescout(city,resource)
     if(need=='loadIn'):
          heli.exeloadIn(city,resource)
     if(need=='woundOut'):
          heli.exewoundOut(city,resource)
     if(need=='equipIn'):
          heli.exeequipIn(city,resource)
     
def findTheWaitHeli(heliList):#找一个现在作业时间最小的飞机，也就是实际仿真中，下一架要调度的飞机
     lTime=100000000000000000
     h=None
     for heli in heliList:
          if(heli.t<lTime):
               lTime=heli.t
               h=heli
     return h

def checkHeliAbility(heli,need):
     judge=False
     if(need=='peopleIn'):
          if(heli.carryPeople>0):
               judge=True
     if(need=='peopleOut'):
          if(heli.carryPeople>0):
               judge=True
     if(need=='scout'):
          if(heli.scoutSpeed>0):
               judge=True
     if(need=='loadIn'):
          if(heli.load>0):
               judge=True
     if(need=='woundOut'):
          if(heli.carryWounded>0):
               judge=True
     if(need=='equipIn'):
          if(heli.equipUpDown>0):
               judge=True
     return judge


def getInfoList():
     wb=openpyxl.load_workbook(plan+".xlsx")
     heliSheet=wb['heli']
     citySheet=wb['city']
     resourceSheet=wb['resource']
     
     
     heliList=[]
     cityList=[]
     resourceList=[]
     
     for i in range(3,heliSheet.max_row+1):
          h=Heli(heliSheet.cell(i,1).value, heliSheet.cell(i,2).value, heliSheet.cell(i,3).value, heliSheet.cell(i,4).value, heliSheet.cell(i,5).value, heliSheet.cell(i,6).value, heliSheet.cell(i,7).value, heliSheet.cell(i,8).value, heliSheet.cell(i,9).value, heliSheet.cell(i,10).value, heliSheet.cell(i,11).value, heliSheet.cell(i,12).value, heliSheet.cell(i,13).value, heliSheet.cell(i,14).value, heliSheet.cell(i,15).value, heliSheet.cell(i,16).value)
          heliList.append(h)
          # print(i)
     
     for i in range(3,citySheet.max_row+1):
          c=City(citySheet.cell(i,1).value, citySheet.cell(i,2).value, citySheet.cell(i,3).value, citySheet.cell(i,4).value, citySheet.cell(i,5).value, citySheet.cell(i,6).value, citySheet.cell(i,7).value, citySheet.cell(i,8).value, citySheet.cell(i,9).value)
          cityList.append(c)
     
     for i in range(3,resourceSheet.max_row+1):
          r=Resource(resourceSheet.cell(i,1).value, resourceSheet.cell(i,2).value, resourceSheet.cell(i,3).value, resourceSheet.cell(i,4).value, resourceSheet.cell(i,5).value, resourceSheet.cell(i,6).value, resourceSheet.cell(i,7).value, resourceSheet.cell(i,8).value, resourceSheet.cell(i,9).value)
          resourceList.append(r)
     return heliList,cityList,resourceList


     
def assignWork(heli):
     #根据所输入的直升机，找到离他最近的城市，考察其需求能不能做，能做就分配给他，全都不能做就换城市；全部城市都没有就销毁这架直升机，如果能做：
     # 最后需要返回任务类型，需求城市和资源城市
     city=None
     need=None
     resource=None
     
     # 做一次初始化，给出第一次的最近城市和任务列表，之后的都在循环中指派和变更
     clist=cityList.copy()
     aCity=findRandomCity(heli.lat, heli.lon, clist)#随机一个城市
     needList=aCity.checkNeeded()
     while(len(needList)==0):
          clist.remove(aCity)
          if(len(clist)==0):#当找不到任何一个城市有任何一个任务可以指派给这个直升机做的时候，就可以销毁这个直升机了
               # 销毁这架飞机,可以在主函数中进行判断和销毁
               return None#结束指派  
          #重新指派城市，并给出需求列表
          aCity=random.choice(clist)
          needList=aCity.checkNeeded()
     
     # while(city==None):#当还没有指派城市的时候进行以下操作：
          
     while(need==None):#当还没有指派任务时进行以下操作/尝试选择该城市的一个任务
          if(len(clist)==0):#当找不到任何一个城市有任何一个任务可以指派给这个直升机做的时候，就可以销毁这个直升机了
               # 销毁这架飞机,可以在主函数中进行判断和销毁
               return None#结束指派     
          
          aNeed=random.choice(needList)
          
          if(checkHeliAbility(heli, aNeed)==True):#这里检查了heli是否能够完成该城市的该任务，如果可以的话指派，不行的话，就换城市
               city=aCity
               need=aNeed
          else:
               needList.remove(aNeed)
          
          while(len(needList)==0):
               clist.remove(aCity)
               if(len(clist)==0):#当找不到任何一个城市有任何一个任务可以指派给这个直升机做的时候，就可以销毁这个直升机了
                    # 销毁这架飞机,可以在主函数中进行判断和销毁
                    return None#结束指派  
               #重新指派城市，并给出需求列表
               aCity=random.choice(clist)
               needList=aCity.checkNeeded()
               # break#在该城市找不到任何一个任务可以指派给该飞机，换个城市
          
          #在一个function中能不能销毁传入的参数对象 ？
     
     #在匹配该需求资源的城市中找到随机的一个
     rlist=findMatchResourceList(need, resourceList)
     resource= findRandomCity(heli.lat, heli.lon, rlist)              
     
     return need,city,resource

def checkAllNeedInCity():
     isEmpty=True
     for city in cityList:
          if(city.peopleIn!=0 or city.peopleOut!=0 or city.scout!=0  or city.loadIn!=0 or city.woundOut!=0 or city.equipIn!=0):
               isEmpty=False
     return isEmpty

def outputLog(loglist):
     loglist.sort(key=lambda x: (x[0], x[-1]))
     with open('.\\plan2\\'+(str(taskMark)+'.txt'),'w',encoding='UTF-8') as f:
          f.write('任务完成总时长:'+str(round(max(realTimeList)/60,3))+' 分钟 \n')
          # f.writelines(taskDoneTime)
          # f.writelines(taskDoneHeli)
          for i in range(len(taskDoneHeli)):
               f.write(("任务"+str(i+1)+' 完成于 '+str(round(taskDoneTime[i]/60,3))+'分钟 by '+taskDoneHeli[i]+'\n'))
          for log in loglist:
               s=''
               for k in log:
                    s+=(str(k)+' ')
               # print(s)
               f.write(s+'\n')
          f.close()
     
          

def handleLoglist(loglist):
     dic={}
     for i in range(len(loglist)):
          if loglist[i][0] in dic.keys():
               dic[loglist[i][0]].append(loglist[i][-1])
          else:
               dic[loglist[i][0]]=[loglist[i][-1]]
     return dic

def draw(a,x):#a是要画的列表，x是在图中的横坐标
# =============================================================================
#      for i in range(len(a)):
#           sums=0
#           for k in range(i):
#                if k>-1:
#                     sums+=a[k]#以上是叠加单个时间段，现在因为是累加和，故简化
# =============================================================================
       for i in range(len(a)):
            plt.bar(x,a[-i-1], bottom=0,width=0.5)

def flowChart(dic):
     plt.figure(figsize=(10,8),dpi=300)
     plt.rcParams['font.sans-serif'] = 'SimHei'#黑体
     plt.rcParams['axes.unicode_minus'] = False
     plt.title('各飞机流程图')
     for key in dic.keys():
          draw(dic[key],key)
     plt.grid(True)
     plt.xticks(rotation=60, horizontalalignment="center")
     plt.xlabel('飞机名称')
     plt.ylabel('time/s')
     # plt.annotate(('任务完成总时长:'+str(max(realTimeList))),(0,0),(0,20000))
     plt.savefig('.\\plan2\\'+str(taskMark)+'.jpg')
     
     
     
if __name__== "__main__":
     loglist=[]
     heliList,cityList,resourceList=getInfoList()
     
     while(checkAllNeedInCity()==False):
          heli=findTheWaitHeli(heliList)#找一个现在作业时间最小的飞机，也就是实际仿真中，下一架要调度的飞机
          
          if(assignWork(heli)!=None):
               need,city,resource=assignWork(heli)
               doOneFunctionByNeeded(heli, need, city, resource)
               # print(heli.name,'飞机作业时间',heli.t)
               
          else:
               # print('该飞机已无任务可做：',heli.name,'飞机作业时间',heli.t)
               realTimeList.append(heli.t)
               heliList.remove(heli)
     for heli in heliList:#版本2更新处，能够直观的看到各直升机的时间参数。
          realTimeList.append(heli.t)
     
     
     
     print('调度完成')
     # print('realtimeList:',realTimeList)
     print('realTime:',max(realTimeList))
     
     taskSettle()#根据保存的数据，清算taskDoneTime和taskDoneHeli
     taskMark=0
     for i in range(9):
          taskMark+=taskWeightList[i]*taskDoneTime[i]/60
     print('taskMark',taskMark)
     print(taskDoneTime,taskDoneHeli)
     
     outputLog(loglist)
     dic=handleLoglist(loglist)
     flowChart(dic)
     
     
     
     
          





