# -*- coding: utf-8 -*-
"""
Created on Tue Mar  9 08:23:03 2021

@author: YU Yixiong
"""
import openpyxl
import static as st
class Heli():
    def __init__(self,para,log):
        self.para=para
        self.log=log
    lon=120.03 
    lat=29.34# presuppose to be in the YiWu Airport
    t=1200 # valued as the first refueling time
    oil=0
    name=''

    # wuzi=0 #goods
    # human=0
    # wg=0# swing
    # yh=0# the space for victims in the calamity
    
    def informBase(self,t1,t2,base):
        # Inform the base that "i will occupy your area between t1-t2, the fleids change based on IsHeli
        if self.para["IsHeli"]==True:
            base.occupyTimeForHeli.append([t1,t2,self.para["HeliArea"]])
        else:
            base.occupyTimeForTrack.append([t1,t2,self.para["TrackArea"]])
    
    def checkPositionForBackBase(self,base,baseList):
         # baseNameList=[]
         # for b in baseList:
         #      baseNameList.append(b.para["Name"])
         if len(self.log)>1:
              if self.log[-1][0]!='基地保障':
                   return True
              elif self.log[-1][1]!=base.para["Name"] and self.log[-2][0]!='基地保障':
                   # allow transfer base for once, but not twice
                   return True
              else:
                   return False
         else:
              return True
              
    def backToBase(self,base):
        d=st.getDistance(self.lat,self.lon,base.para["PosY"],base.para["PosX"])
        self.t+=d/self.para["Speed"]*3600
        self.log.append(['前往基地', base.para["Name"],'',self.t])

        t1=self.t

        self.t+=1200 
        t2=self.t
        self.oil=self.para["OilMax"]
        self.lat=base.para["PosY"]
        self.lon=base.para["PosX"]
        
        self.log.append(['基地保障', base.para["Name"],'20分钟',self.t])
        self.informBase(t1, t2, base)

    def modifyLog(self):
         dl=[]
         for i in range(-1,-len(self.log)-1,-1):
              if self.log[i][0]!="前往基地" or self.log[i][0]!="基地保障":
                   # return None
                   break
              else: 
                   dl.append(self.log[i])
         for l in dl:
              self.log.remove(l)
    # def exepeopleIn(self,c1,c2):
        # define this mission by phase and you will get the time c0--c1--c2 anddo task in c1, c2
        # use the return to confirm the mission enforceability

def addHeli(name):
# def addHeli(name,city):
    wb=openpyxl.load_workbook("../data/SystemParameter.xlsx")
    sheet=wb['飞行器']
    
    h=Heli({},[])
    for i  in range(3,15):
        if sheet.cell(i,2).value==name:
            for j in range(1,sheet.max_column+1):    
                h.para[sheet.cell(1,j).value]=sheet.cell(i,j).value   
    wb.close()
    
    # initalize the temp variables from the
    h.oil=h.para["OilMax"]
    # h.lat=city.para["PosY"]
    # h.lon=city.para["PosX"]   
    return h

def writeHeliLog(h,path):
     print(h.name+"write heli log")
     with open(path,'a+') as f:
          for l in h.log:
              f.write(h.name+':'+' '+str(l[0])+' '+str(l[1])+' '+str(l[2])+'   t='+str(l[3])+'\n')
          f.write('\n\n')
          f.close()


if __name__ == "__main__":
    pass
    # p=addHeli('AC313_YL',)
# =============================================================================
# verify the indenpendence of each instance
# c=Heli()
# print(c.lon)
# c.lon=10
# d=Heli()
# print(d.lon)
# =============================================================================
