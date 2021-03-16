# -*- coding: utf-8 -*-
"""
Created on Tue Mar  9 14:49:07 2021

@author: YU Yixiong
"""

import openpyxl
import math
class City():
    def  __init__(self,para,trackNum,heliNum,occupyTimeForHeli,occupyTimeForTrack):
        self.para=para
        self.trackNum=trackNum
        self.heliNum=heliNum
        self.occupyTimeForHeli=occupyTimeForHeli
        self.occupyTimeForTrack=occupyTimeForTrack
    

    def checkHeliNum(self,t1,t2,area):
        #the occupatin list is composed with [t1,t2,area]
        if len(self.occupyTimeForHeli)==0:
             return True
        
        t1=math.ceil(t1)
        t2=math.floor(t2)
        for t in range(t1,t2,10):
             heliNum=self.para["MaxHeliNum"]
             for tup in self.occupyTimeForHeli:
                 if ( t>tup[0] and t<tup[1] ):
                     heliNum-=tup[2]
             if area>heliNum :
                  return False
             # print(self.para["Name"],"剩余直升机停机面积：",heliNum)
        return(True)
    # this is a place showing the anti-AI of this system. The heli won't wait for a leisure, hovering there.
    
    def checkTrackNum(self,t1,t2,area):
        if len(self.occupyTimeForTrack)==0:
             return True
        t1=math.ceil(t1)
        t2=math.floor(t2)
        for t in range(t1,t2,10):
             trackNum=self.para["MaxTrackNum"]
             for tup in self.occupyTimeForTrack:
                 if ( t>tup[0] and t<tup[1] ):
                     trackNum-=tup[2]
             if area>trackNum:
                  return False
             # print(self.para["Name"],"剩余固定翼停机面积：",trackNum)
        return(True)
    
# =============================================================================
#     def checkLandingFeasibilityForHeli(self):
#         if self.pheliNum
# =============================================================================

def getCityList():
    wb=openpyxl.load_workbook("../data/SystemParameter.xlsx")
    sheet=wb['地点']
    cityList=[]
    
    for i in range(3,15):
        c=City({},0,0,[],[])
        #字典被append给列表的是一个地址，it will be update at the end 
        for j in range(1,sheet.max_column+1):    
            c.para[sheet.cell(1,j).value]=sheet.cell(i,j).value
        c.trackNum=c.para["MaxTrackNum"]
        c.heliNum=c.para["MaxHeliNum"]
        cityList.append(c)
        
    wb.close()
    return cityList

def addCity(r):#create a city by the row index of the sheet
    wb=openpyxl.load_workbook("../data/SystemParameter.xlsx")
    sheet=wb['地点']
    c=City({},0,0,[],[])
    for j in range(1,sheet.max_column+1):    
            c.para[sheet.cell(1,j).value]=sheet.cell(r,j).value
    # print(id(c.para))
    c.trackNum=c.para["MaxTrackNum"]
    c.heliNum=c.para["MaxHeliNum"]
    wb.close()
    
    return c

if __name__ == "__main__":
    # cl=[]
    # for i in range(3,15):
    #     c=addCity(i)
    #     cl.append(copy.deepcopy( c ) )
    # cl=getCityList()
    pass