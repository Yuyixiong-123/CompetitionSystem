# -*- coding: utf-8 -*-
"""
Created on Tue Mar  9 16:35:31 2021

@author: YU Yixiong
"""

import openpyxl
import commander
import json

class Mission:
    def __init__(self,para,residue,log,seri):
        self.para=para
        self.residue=residue
        self.log=log
        self.seri=seri
        
    def residueCheck(self):
        '''
        return the need of mission name

        '''
        for key in self.residue.keys():
            if self.residue[key]>0:
                return key
        return None                
            
def getMissionList():
    wb=openpyxl.load_workbook("../data/SystemParameter.xlsx")
    sheet=wb['任务']
    missionList=[]
    
    for i in range(3,12):
        m=Mission({},{},[],i-2)
        for j in range(1,sheet.max_column+1):    
            m.para[sheet.cell(1,j).value]=sheet.cell(i,j).value
            if(1<j and j<9):
                m.residue[sheet.cell(1,j).value]=sheet.cell(i,j).value
            
        missionList.append(m)
        
    return missionList

def writeMissionLog(m):
     for i in range(len(m.log)):
          m.log[i].insert(0,m.seri)
     orilog=[]
     with open(commander.config.missionLogPath,'r') as f:
          orilog=json.load(f)
          f.close()
     with open(commander.config.missionLogPath,'w') as f:
          json.dump(orilog+m.log,f)
          f.close()
          
if __name__ =="__main__":
    # p=getMissionList()
    pass